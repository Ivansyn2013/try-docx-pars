from openpyxl import *
from openpyxl.utils import *
import re

RE_DIG_F = re.compile(r'\d{2}\.\d{2}\.') # патерн для цифр
RE_ALB = re.compile(r'\w{1}\.') #патерн для первой буквы

# берем фаил
start_file = r'G:\Python project\try-docx-pars\sample_files\testord-2.xlsx'

wb = load_workbook(filename= start_file)
sheet_ranges = wb['questions']
#print(sheet_ranges['A1'].value)

ws = wb['questions'] #Получить лист с именем ...

rows = ws.rows
col = ws.columns



#print(ws.cell(row=1, column=5).value)
#print (sheet_ranges.max_row) # число строк ( примрно также число стобцов) через объект листа
i=1
while i <= ws.max_row :
    #убираем цифры
    ws.cell(row=i, column=3).value = re.sub(RE_DIG_F, '', str(ws.cell(row=i, column=3).value))
    #убираем буквы
    for li in range(4,9):
        ws.cell(row=i, column=li).value = re.sub(RE_ALB,'',str(ws.cell(row=i, column=li).value))

    f_alb = str(ws.cell(row=i, column=9).value)
    if f_alb.lower().find('а') != -1 :
        ws.cell(row=i, column=4).value += '*'
    elif f_alb.lower().find('б') != -1:
        ws.cell(row=i, column=5).value += '*'
    elif f_alb.lower().find('в') != -1:
        ws.cell(row=i, column=6).value += '*'
    elif f_alb.lower().find('г') != -1:
        ws.cell(row=i, column=7).value += '*'
    elif f_alb.lower().find('д') != -1:
        ws.cell(row=i, column=8).value += '*'


    i+=1


for row in rows:
    line = [coll.value for coll in row]



res_filename = '2result.xlsx'
wb.save(filename=res_filename)